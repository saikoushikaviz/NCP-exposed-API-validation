"""
NCP API Validation - Excel Report Generator

Generates a styled Excel report with:
  - Sheet 1: Summary
  - Sheet 2: Functional Flow Tests
"""

import os
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colour palette ────────────────────────────────────────────
GREEN_FILL  = PatternFill("solid", fgColor="C6EFCE")
RED_FILL    = PatternFill("solid", fgColor="FFC7CE")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
ALT_FILL    = PatternFill("solid", fgColor="EEF3FB")

GREEN_FONT  = Font(color="276221", bold=True)
RED_FONT    = Font(color="9C0006", bold=True)
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT  = Font(bold=True, size=13)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

THIN   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ── Public API ────────────────────────────────────────────────

class ReportCollector:
    """Collects test results during the pytest run."""

    def __init__(self):
        self.flow_rows: list[dict] = []

    def add_flow(self, *, step, description, api_method, endpoint,
                 expected_status, actual_status, response_summary, passed: bool):
        self.flow_rows.append({
            "step":             step,
            "description":      description,
            "api_method":       api_method,
            "endpoint":         endpoint,
            "expected_status":  expected_status,
            "actual_status":    actual_status,
            "response_summary": response_summary,
            "passed":           passed,
        })


def generate_report(collector: ReportCollector, output_dir: str = ".") -> str:
    """Write the Excel report and return the file path."""
    path = os.path.join(output_dir, "NCP_Exposed_API_Report.xlsx")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _build_flow_sheet(wb, collector.flow_rows)
    _build_summary_sheet(wb, collector)

    wb.save(path)
    return path


# ── Private helpers ───────────────────────────────────────────

def _apply_header(ws, headers: list[str], col_widths: list[int]):
    ws.append(headers)
    for col_idx, (_, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTER
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22


def _result_cell(ws, row: int, col: int, passed: bool):
    cell = ws.cell(row=row, column=col, value="PASS" if passed else "FAIL")
    cell.fill      = GREEN_FILL if passed else RED_FILL
    cell.font      = GREEN_FONT if passed else RED_FONT
    cell.alignment = CENTER
    cell.border    = BORDER


def _data_cell(ws, row: int, col: int, value, alt: bool = False):
    cell = ws.cell(row=row, column=col, value=str(value) if value is not None else "")
    cell.alignment = LEFT
    cell.border    = BORDER
    if alt:
        cell.fill = ALT_FILL


def _build_flow_sheet(wb, rows: list[dict]):
    ws = wb.create_sheet("Functional Flow")

    headers    = ["Step", "Description", "Method", "Endpoint",
                  "Expected Status", "Actual Status", "Response Body (JSON)", "Result"]
    col_widths = [6,      32,            10,        45,
                  17,               15,             70,                        8]
    _apply_header(ws, headers, col_widths)

    ws.freeze_panes = "A2"

    for i, r in enumerate(rows, start=1):
        row_num = i + 1
        alt     = (i % 2 == 0)
        _data_cell(ws, row_num, 1, r["step"],             alt)
        _data_cell(ws, row_num, 2, r["description"],      alt)
        _data_cell(ws, row_num, 3, r["api_method"],       alt)
        _data_cell(ws, row_num, 4, r["endpoint"],         alt)
        _data_cell(ws, row_num, 5, r["expected_status"],  alt)
        _data_cell(ws, row_num, 6, r["actual_status"],    alt)
        _data_cell(ws, row_num, 7, r["response_summary"], alt)
        _result_cell(ws, row_num, 8, r["passed"])
        lines = r["response_summary"].count("\n") + 1
        ws.row_dimensions[row_num].height = max(18, min(lines * 13, 300))


def _build_summary_sheet(wb, collector: ReportCollector):
    ws = wb.create_sheet("Summary", 0)

    flow_total  = len(collector.flow_rows)
    flow_passed = sum(1 for r in collector.flow_rows if r["passed"])
    flow_failed = flow_total - flow_passed

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    def _title(row, text):
        c = ws.cell(row=row, column=1, value=text)
        c.font      = TITLE_FONT
        c.alignment = LEFT

    def _row(row, label, value, highlight=False):
        lc = ws.cell(row=row, column=1, value=label)
        vc = ws.cell(row=row, column=2, value=value)
        lc.alignment = LEFT
        vc.alignment = CENTER
        lc.border    = BORDER
        vc.border    = BORDER
        if highlight:
            fill = GREEN_FILL if (isinstance(value, int) and value == 0) \
                              or value == "ALL PASSED" else RED_FILL
            vc.fill = fill

    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    _title(1, "NCP Exposed API — Validation Report")
    ws.cell(row=2, column=1, value=f"Generated: {ts}").font = Font(italic=True, color="666666")

    _title(4, "Functional Flow Tests")
    _row(5, "Total",  flow_total)
    _row(6, "Passed", flow_passed)
    _row(7, "Failed", flow_failed, highlight=True)

    _title(9, "Overall")
    _row(10, "Total Tests", flow_total)
    _row(11, "Passed",      flow_passed)
    _row(12, "Failed",      flow_failed, highlight=True)
    _row(13, "Result",
         "ALL PASSED" if flow_failed == 0 else f"{flow_failed} FAILED",
         highlight=True)
